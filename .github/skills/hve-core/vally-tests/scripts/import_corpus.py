#!/usr/bin/env python3
# Copyright (c) Microsoft Corporation.
# SPDX-License-Identifier: MIT
"""Vally corpus importer.

Reads a CSV or XLSX file matching the canonical column contract, validates each
row, normalizes prompts (trim + lowercase + collapse-whitespace), hashes them
with SHA-256, runs the skill-local safety lint as a per-row subprocess, dedupes
against an optional target eval YAML, and emits both an append-only YAML patch
and a JSON report.  Every accepted row is forced to ``tags.advisory: true``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import sys
import tempfile
import unicodedata
from collections.abc import Iterable, Iterator
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import yaml.reader

REQUIRED_COLUMNS: tuple[str, ...] = (
    "prompt",
    "kind",
    "target_artifact",
    "grader",
    "tags",
    "expected_refusal_category",
    "notes",
)

ALLOWED_KINDS: frozenset[str] = frozenset({"agent", "prompt", "instructions", "skill"})

EXISTING_HASH_RE = re.compile(r"#\s*sha256:([0-9a-f]{64})", re.IGNORECASE)
SAFETY_CATEGORY_RE = re.compile(r"category=(\S+)")


class CorpusImportError(RuntimeError):
    """Raised on irrecoverable corpus-import errors (caller exits non-zero)."""


@dataclass
class ImportReport:
    source: str
    target: str | None
    patch_path: str
    timestamp: str
    accepted: list[dict[str, object]] = field(default_factory=list)
    rejected: list[dict[str, object]] = field(default_factory=list)
    flagged: list[dict[str, object]] = field(default_factory=list)
    duplicates: list[dict[str, object]] = field(default_factory=list)

    def totals(self) -> dict[str, int]:
        return {
            "accepted": len(self.accepted),
            "rejected": len(self.rejected),
            "flagged": len(self.flagged),
            "duplicates": len(self.duplicates),
        }

    def to_dict(self) -> dict[str, object]:
        return {
            "source": self.source,
            "target": self.target,
            "patch_path": self.patch_path,
            "timestamp": self.timestamp,
            "totals": self.totals(),
            "accepted": self.accepted,
            "rejected": self.rejected,
            "flagged": self.flagged,
            "duplicates": self.duplicates,
        }


def normalize_prompt(value: str) -> str:
    """Apply NFC + trim + lowercase + whitespace-collapse for dedupe hashing."""
    if value is None:
        return ""
    nfc = unicodedata.normalize("NFC", str(value))
    return re.sub(r"\s+", " ", nfc.strip().lower())


def hash_prompt(normalized: str) -> str:
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def _strip(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_csv_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise CorpusImportError(f"{path}: no header row")
        missing = [c for c in REQUIRED_COLUMNS if c not in reader.fieldnames]
        if missing:
            raise CorpusImportError(
                f"{path}: missing required columns: {', '.join(missing)}"
            )
        for row in reader:
            yield {k: _strip(v) for k, v in row.items()}


def read_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as exc:  # pragma: no cover
        raise CorpusImportError("openpyxl is required to import .xlsx files") from exc
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise CorpusImportError(f"{path}: workbook has no active sheet")
    header_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(header_iter, None)
    if header_row is None:
        raise CorpusImportError(f"{path}: no header row")
    headers = [_strip(cell) for cell in header_row]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise CorpusImportError(
            f"{path}: missing required columns: {', '.join(missing)}"
        )
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        yield {
            headers[index]: _strip(cell)
            for index, cell in enumerate(row)
            if index < len(headers) and headers[index]
        }


def read_rows(path: Path) -> Iterator[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path)
    raise CorpusImportError(f"{path}: unsupported suffix '{suffix}'; use .csv or .xlsx")


def validate_row(row: dict[str, str], line_no: int) -> str | None:
    if not row.get("prompt"):
        return f"row {line_no}: empty prompt"
    if not row.get("target_artifact"):
        return f"row {line_no}: empty target_artifact"
    kind = row.get("kind", "")
    if kind not in ALLOWED_KINDS:
        return f"row {line_no}: kind '{kind}' not in {sorted(ALLOWED_KINDS)}"
    return None


def safety_check(
    prompt: str,
    lint_script: Path,
    *,
    pwsh: str = "pwsh",
    timeout_seconds: float = 60.0,
) -> dict[str, object]:
    """Run the skill-local safety lint against a single prompt string."""
    if not lint_script.exists():
        return {
            "exit_code": -1,
            "output": f"safety lint not found: {lint_script}",
            "category": None,
        }
    tmp = tempfile.NamedTemporaryFile(
        "w", suffix=".txt", delete=False, encoding="utf-8"
    )
    try:
        tmp.write(prompt)
        tmp.close()
        tmp_path = Path(tmp.name)
        try:
            result = subprocess.run(
                [pwsh, "-NoProfile", "-File", str(lint_script), str(tmp_path)],
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
                check=False,
            )
        except FileNotFoundError:
            return {
                "exit_code": -1,
                "output": f"executable not found: {pwsh}",
                "category": None,
            }
        except subprocess.TimeoutExpired as exc:
            return {
                "exit_code": -1,
                "output": f"safety lint timed out after {exc.timeout}s",
                "category": None,
            }
    finally:
        try:
            Path(tmp.name).unlink(missing_ok=True)
        except OSError:
            pass
    output = ((result.stdout or "") + (result.stderr or "")).strip()
    match = SAFETY_CATEGORY_RE.search(output)
    return {
        "exit_code": result.returncode,
        "output": output,
        "category": match.group(1) if match else None,
    }


def load_existing_hashes(target_path: Path | None) -> set[str]:
    if target_path is None or not target_path.exists():
        return set()
    text = target_path.read_text(encoding="utf-8")
    return {match.group(1).lower() for match in EXISTING_HASH_RE.finditer(text)}


def _indent_block(text: str, prefix: str) -> str:
    lines = text.splitlines() or [""]
    return "".join(f"{prefix}{line}\n" for line in lines)


# Characters PyYAML rejects when reading a stream. Reuse the reader's own
# NON_PRINTABLE pattern so this matches PyYAML exactly: a prompt containing any
# of these bytes cannot ride in a literal block scalar and must be emitted as a
# double-quoted scalar instead.
_YAML_NON_PRINTABLE = yaml.reader.Reader.NON_PRINTABLE


def _yaml_scalar(value: str) -> str:
    """Render a string as a safely-quoted YAML scalar.

    ``json.dumps`` emits a double-quoted form whose quoting and escaping are
    valid YAML, so YAML-significant characters (``:``, ``#``, leading ``-``,
    quotes, embedded newlines) cannot corrupt the surrounding document.
    """
    return json.dumps(value)


def _comment_value(value: str) -> str:
    """Collapse line breaks so an interpolated value stays on one comment line.

    A bare newline in a YAML comment terminates the comment, so unsanitized
    values could inject document content. Replacing line breaks with spaces
    keeps the comment single-line and inert.
    """
    return value.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")


def _block_scalar_safe(prompt: str) -> bool:
    """Return True when ``prompt`` can ride safely in a literal block scalar.

    A literal block scalar auto-detects its content indentation from the first
    non-empty line. A leading empty line, or a first line carrying leading
    whitespace, makes that detection ambiguous: a later, less-indented line
    de-indents below the detected level, terminates the scalar early, and
    corrupts the surrounding block mapping. Control characters cannot survive a
    literal block scalar at all. Such prompts must use a double-quoted scalar.
    """
    if _YAML_NON_PRINTABLE.search(prompt):
        return False
    lines = prompt.splitlines()
    if not lines:
        return False
    first = lines[0]
    return first != "" and first[0] not in " \t"


def build_patch_entry(row: dict[str, str], digest: str) -> str:
    prompt = row["prompt"]
    if _block_scalar_safe(prompt):
        prompt_lines = ["- prompt: |\n", _indent_block(prompt, "    ")]
    else:
        # Fall back to a double-quoted scalar so the patch round-trips through
        # yaml.safe_load even when a literal block scalar would be ambiguous.
        prompt_lines = [f"- prompt: {_yaml_scalar(prompt)}\n"]
    parts: list[str] = [
        f"# sha256:{digest}\n",
        f"# kind:{_comment_value(row['kind'])}\n",
        f"# target:{_comment_value(row['target_artifact'])}\n",
        *prompt_lines,
        f"  grader: {_yaml_scalar(row['grader'] or '<unset>')}\n",
        "  tags:\n",
    ]
    raw_tags = row.get("tags", "")
    if raw_tags:
        parts.append(f"    raw: {_yaml_scalar(raw_tags)}\n")
    parts.append("    advisory: true\n")
    expected = row.get("expected_refusal_category", "")
    if expected:
        parts.append(f"  expected_refusal_category: {_yaml_scalar(expected)}\n")
    notes = row.get("notes", "")
    if notes:
        parts.append(f"  notes: {_yaml_scalar(notes)}\n")
    return "".join(parts)


def import_corpus(
    source: Path,
    *,
    target: Path | None = None,
    report_dir: Path,
    lint_script: Path,
    skip_safety: bool = False,
    pwsh: str = "pwsh",
    now: datetime | None = None,
) -> tuple[ImportReport, Path, Path]:
    if not source.exists():
        raise CorpusImportError(f"source file not found: {source}")
    report_dir.mkdir(parents=True, exist_ok=True)
    timestamp = (now or datetime.now(timezone.utc)).strftime("%Y%m%dT%H%M%SZ")
    report_path = report_dir / f"vally-test-author-import-{timestamp}.json"
    patch_path = report_dir / f"vally-test-author-import-{timestamp}.patch.yml"

    report = ImportReport(
        source=str(source),
        target=str(target) if target else None,
        patch_path=str(patch_path),
        timestamp=timestamp,
    )

    existing_hashes = load_existing_hashes(target)
    accepted_blocks: list[str] = []

    rows: Iterable[dict[str, str]] = read_rows(source)
    for index, row in enumerate(rows, start=2):  # header is line 1
        error = validate_row(row, index)
        if error:
            report.rejected.append({"line": index, "reason": error, "row": row})
            continue

        normalized = normalize_prompt(row["prompt"])
        digest = hash_prompt(normalized)
        if digest in existing_hashes:
            report.duplicates.append({"line": index, "sha256": digest, "row": row})
            continue
        existing_hashes.add(digest)

        if not skip_safety:
            safety = safety_check(row["prompt"], lint_script, pwsh=pwsh)
            if safety["exit_code"] != 0:
                report.flagged.append({"line": index, "safety": safety, "row": row})
                continue

        accepted_blocks.append(build_patch_entry(row, digest))
        report.accepted.append({"line": index, "sha256": digest, "row": row})

    patch_path.write_text("\n".join(accepted_blocks), encoding="utf-8")
    report_path.write_text(
        json.dumps(report.to_dict(), indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    return report, report_path, patch_path


def _default_lint_script() -> Path:
    return Path(__file__).resolve().parent / "Lint-VallyTestSafety.ps1"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="import_corpus",
        description=(
            "Import a Vally corpus CSV/XLSX into an append-only patch "
            "with safety + dedupe gating."
        ),
    )
    parser.add_argument(
        "source",
        help="CSV or XLSX source file matching the canonical column contract.",
    )
    parser.add_argument(
        "--target",
        default=None,
        help="Existing eval YAML for dedupe comparison (optional).",
    )
    parser.add_argument(
        "--report-dir",
        default="logs",
        help="Directory for JSON report + patch output. Default: logs/.",
    )
    parser.add_argument(
        "--lint-script",
        default=None,
        help=(
            "Override path to the skill-local Lint-VallyTestSafety.ps1. "
            "Default: sibling script next to import_corpus.py."
        ),
    )
    parser.add_argument(
        "--pwsh",
        default="pwsh",
        help="Executable to invoke for the safety lint. Default: pwsh.",
    )
    parser.add_argument(
        "--skip-safety",
        action="store_true",
        help=(
            "Skip the per-row safety lint subprocess (for offline test environments)."
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    source = Path(args.source).resolve()
    target = Path(args.target).resolve() if args.target else None
    report_dir = Path(args.report_dir).resolve()
    lint_script = (
        Path(args.lint_script).resolve() if args.lint_script else _default_lint_script()
    )
    try:
        report, report_path, patch_path = import_corpus(
            source,
            target=target,
            report_dir=report_dir,
            lint_script=lint_script,
            skip_safety=args.skip_safety,
            pwsh=args.pwsh,
        )
    except CorpusImportError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    totals = report.totals()
    print(f"report: {report_path}")
    print(f"patch:  {patch_path}")
    print(" ".join(f"{key}={value}" for key, value in totals.items()))
    return 0 if (totals["rejected"] == 0 and totals["flagged"] == 0) else 1


if __name__ == "__main__":
    sys.exit(main())
